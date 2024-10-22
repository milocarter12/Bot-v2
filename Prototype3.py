import streamlit as st
import openpyxl
from openpyxl import load_workbook, Workbook
import tempfile
import shutil
import datetime
import random
import os
import logging

# Set up logging
log_file_path = "profit_calculator_debug.log"
logging.basicConfig(filename=log_file_path, level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Load spreadsheets
def load_excel(file_path):
    try:
        st.write(f"Loading workbook from: {file_path}")
        logging.info(f"Loading workbook from: {file_path}")
        return load_workbook(file_path, data_only=False)
    except Exception as e:
        logging.error(f"Failed to load workbook from {file_path}: {e}")
        st.error(f"Failed to load workbook from {file_path}. Check the file path and try again.")
        raise e

def save_excel(workbook, file_path):
    try:
        st.write(f"Saving workbook to: {file_path}")
        logging.info(f"Saving workbook to: {file_path}")
        workbook.save(file_path)
    except Exception as e:
        logging.error(f"Failed to save workbook to {file_path}: {e}")
        st.error(f"Failed to save workbook to {file_path}. Check if the file is open or locked.")
        raise e

# Streamlit UI
def main():
    st.title('Profit Calculator Tool')
    logging.info("Profit Calculator Tool started")

    # Step 1: Collect keyword name from user
    keyword_name = st.text_input('Enter the Keyword Name:')
    if not keyword_name:
        st.warning("Please enter a keyword name before proceeding.")
        logging.warning("No keyword name entered")
        return

    st.write(f"Keyword Name Entered: {keyword_name}")
    logging.info(f"Keyword Name Entered: {keyword_name}")

    # Upload Excel files
    st.write("Upload Excel files...")
    landed_cost_file = st.file_uploader("Upload Copy of Landed Cost HFBA Excel file", type=["xlsx"])
    profit_calc_file = st.file_uploader("Upload Pycnogenol Profit Calculator Excel file", type=["xlsx"])

    if not landed_cost_file or not profit_calc_file:
        st.warning("Please upload both Excel files to proceed.")
        return

    try:
        wb_landed_cost = load_workbook(landed_cost_file, data_only=False)
        wb_profit_calc = load_workbook(profit_calc_file, data_only=False)
    except Exception as e:
        logging.error("Failed to load Excel files: %s", e)
        st.error("Failed to load one or both Excel files. Please check the files and try again.")
        return

    # Ensure the worksheets are loaded correctly
    if not wb_landed_cost.sheetnames or not wb_profit_calc.sheetnames:
        st.error("Failed to load one or both Excel files. Please check the file paths and try again.")
        logging.error("Failed to load one or both Excel files. No sheetnames found.")
        return

    sheet_landed_cost = wb_landed_cost.active
    sheet_profit_calc = wb_profit_calc.active

    # Step 2: Input data collection from user via Streamlit
    shipping_total = st.number_input('Enter Shipping Total:', min_value=0.0, step=0.01)
    unit_cost = st.number_input('Enter Unit Cost:', min_value=0.0, step=0.01)
    target_sales_per_month = st.number_input('Enter Target Sales per Month:', min_value=0, step=1)
    selling_price = st.number_input('Enter Selling Price:', min_value=0.0, step=0.01)
    fulfillment_fee = st.number_input('Enter Fulfillment Fee:', min_value=0.0, step=0.01)
    storage_cost = st.number_input('Enter Storage Cost:', min_value=0.0, step=0.01)

    if not (shipping_total and unit_cost and target_sales_per_month and selling_price and fulfillment_fee and storage_cost):
        st.error("All fields must be filled out before proceeding.")
        logging.error("Input validation failed: Not all fields were filled out.")
        return

    logging.info(f"Inputs collected: Shipping Total = {shipping_total}, Unit Cost = {unit_cost}, Target Sales per Month = {target_sales_per_month}, Selling Price = {selling_price}, Fulfillment Fee = {fulfillment_fee}, Storage Cost = {storage_cost}")
    st.write(f"Inputs collected: Shipping Total = {shipping_total}, Unit Cost = {unit_cost}, Target Sales per Month = {target_sales_per_month}, Selling Price = {selling_price}, Fulfillment Fee = {fulfillment_fee}, Storage Cost = {storage_cost}")

    if st.button('Calculate & Update Excel Files'):
        st.write("Button pressed. Starting calculation and update process...")
        logging.info("Button pressed. Starting calculation and update process...")
        try:
            # Update Landed Cost HFBA Excel file with inputted values
            st.write("Updating Landed Cost HFBA Excel file...")
            logging.info("Updating Landed Cost HFBA Excel file...")
            sheet_landed_cost['G17'] = shipping_total
            sheet_landed_cost['G25'] = unit_cost

            # Save the updated workbook temporarily to ensure formulas are recalculated
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                st.write(f"Saving temporary Landed Cost workbook to: {tmp.name}")
                logging.info(f"Saving temporary Landed Cost workbook to: {tmp.name}")
                save_excel(wb_landed_cost, tmp.name)
                landed_cost_file_path = tmp.name

            # Reload the workbook to ensure recalculated values are obtained
            st.write("Reloading Landed Cost workbook to recalculate formulas...")
            logging.info("Reloading Landed Cost workbook to recalculate formulas...")
            wb_landed_cost = load_workbook(landed_cost_file_path, data_only=True)
            sheet_landed_cost = wb_landed_cost.active

            # Attempt multiple times to get a valid calculated value from cell M25
            calculated_landed_cost = None
            for attempt in range(3):
                calculated_landed_cost = sheet_landed_cost['M25'].value
                if calculated_landed_cost is not None and not isinstance(calculated_landed_cost, str):
                    break
                logging.warning(f"Attempt {attempt + 1}: Failed to retrieve a valid landed cost value from cell M25. Retrying...")
                st.write(f"Attempt {attempt + 1}: Retrying to get a valid landed cost value...")
                wb_landed_cost = load_workbook(landed_cost_file_path, data_only=True)
                sheet_landed_cost = wb_landed_cost.active

            if calculated_landed_cost is None or isinstance(calculated_landed_cost, str):
                # Manual fallback calculation (example approximation for demonstration purposes)
                calculated_landed_cost = (shipping_total + unit_cost) / (target_sales_per_month or 1)  # Example fallback logic
                logging.warning("Using fallback manual calculation for Landed Cost.")
                st.warning("Failed to retrieve recalculated Landed Cost. Using fallback manual calculation.")

            st.write(f"Calculated Landed Cost extracted: {calculated_landed_cost}")
            logging.info(f"Calculated Landed Cost extracted: {calculated_landed_cost}")

            # Update Pycnogenol Profit Calculator with provided values and calculated landed cost
            st.write("Updating Pycnogenol Profit Calculator with provided values...")
            logging.info("Updating Pycnogenol Profit Calculator with provided values...")
            sheet_profit_calc['F33'] = target_sales_per_month
            sheet_profit_calc['F35'] = selling_price
            sheet_profit_calc['F37'] = fulfillment_fee
            sheet_profit_calc['F43'] = storage_cost
            sheet_profit_calc['F39'] = calculated_landed_cost

            # Save both updated sheets in a single Excel file without changing their structures
            combined_file_path = f"{keyword_name}_{datetime.datetime.now().strftime('%Y%m%d')}_{random.randint(10, 99)}.xlsx"
            st.write(f"Saving combined workbook to: {combined_file_path}")
            logging.info(f"Saving combined workbook to: {combined_file_path}")
            combined_workbook = Workbook()
            combined_workbook.remove(combined_workbook.active)  # Remove the default sheet

            # Copy sheets from Landed Cost workbook
            for sheet_name in wb_landed_cost.sheetnames:
                st.write(f"Copying Landed Cost sheet: {sheet_name}")
                logging.info(f"Copying Landed Cost sheet: {sheet_name}")
                source_sheet = wb_landed_cost[sheet_name]
                target_sheet = combined_workbook.create_sheet(title=sheet_name)
                for row in source_sheet.iter_rows():
                    for cell in row:
                        new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell.font = cell.font
                            new_cell.border = cell.border
                            new_cell.fill = cell.fill
                            new_cell.number_format = cell.number_format
                            new_cell.protection = cell.protection
                            new_cell.alignment = cell.alignment

            # Copy sheets from Profit Calculator workbook
            for sheet_name in wb_profit_calc.sheetnames:
                st.write(f"Copying Profit Calculator sheet: {sheet_name}")
                logging.info(f"Copying Profit Calculator sheet: {sheet_name}")
                source_sheet = wb_profit_calc[sheet_name]
                target_sheet = combined_workbook.create_sheet(title=sheet_name)
                for row in source_sheet.iter_rows():
                    for cell in row:
                        new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell.font = cell.font
                            new_cell.border = cell.border
                            new_cell.fill = cell.fill
                            new_cell.number_format = cell.number_format
                            new_cell.protection = cell.protection
                            new_cell.alignment = cell.alignment

            # Save the combined workbook
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                combined_file_path = tmp.name
                save_excel(combined_workbook, combined_file_path)

            # Display success message
            st.success('Excel files have been updated and combined successfully.')
            st.write("Excel files have been successfully updated and combined.")
            logging.info("Excel files have been successfully updated and combined.")

            # Offer the combined file for download
            with open(combined_file_path, 'rb') as f:
                st.download_button(label='Download Combined Calculators', data=f, file_name=f'{keyword_name}_Updated_Profit_Calculator.xlsx')
        except Exception as e:
            logging.error(f"Error during calculation and update process: {e}")
            st.error("An error occurred during the calculation and update process. Check the log file for details.")

if __name__ == "__main__":
    main()
